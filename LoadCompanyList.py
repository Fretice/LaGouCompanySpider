import os
import json
import requests
import datetime
from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl import load_workbook

#获取城市ID列表
def get_cityId_list(url):
    city_list = []
    html = pq(url= url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        aId = pq(areaId).attr('href').replace('http://www.lagou.com/gongsi/', '').replace('-0-0#filterBox', '')
        if(aId=='0'):
            continue
        city_list.append(aId)

    return city_list
#获取城市名称列表
def get_city_name_list(u):
    city_name_list = []
    url = 'http://www.lagou.com/gongsi/'
    html = pq(url=url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        area_name=pq(areaId).html()
        if area_name=="全国":
            continue
        city_name_list.append(area_name)
    return city_name_list

#获取城市下一共有多少页
def get_city_page(areaId,page_num):
    try:
	    param = {'first': 'false', 'pn': page_num, 'sortField': '0', 'havemark': '0'} #访问参数
	    r = requests.post('http://www.lagou.com/gongsi/'+areaId+'-0-0.json',params=param ) #requsets请求
	    page_num += 1
	    if(len(r.json()['result'])/16==1):
		    return get_city_page(areaId,page_num)
	    else:
		    return page_num
    except:
        return page_num-1

#根据城市ID获取所有公司信息
def get_company_list(areaId):
    company_list = []
    city_page_total=get_city_page(areaId,1)
    for pageIndex in range(1,city_page_total):
        print('正在爬取第'+str(pageIndex)+'页')
        json_url = 'http://www.lagou.com/gongsi/'+areaId+'-0-0.json'
        param = {'first': 'false', 'pn': str(pageIndex), 'sortField': '0', 'havemark': '0'} #访问参数
        r = requests.post(json_url,params=param ) #requsets请求
        msg = json.loads(r.text)
        try:
            for company in msg['result']:
                company_list.append([company['city'],company['cityScore'],company['companyFeatures'],company['companyId'],company['companyLabels'],company['companyLogo'],company['companyName'],str(company['companyPositions']),company['companyShortName'],company['countryScore'],company['createTime'],company['finaceStage'],company['industryField'],company['interviewRemarkNum'],company['otherLabels'], company['positionNum'],company['processRate'],str(datetime.datetime.now())])
        except:
            print('爬取编号为'+str(areaId)+'城市时第'+str(pageIndex)+'页出现了错误,错误时请求返回内容为：'+str(msg))
            continue
    return company_list
#写入Excel文件方法
def write_file(fileName):
    list = []
    wb = Workbook()
    ws = wb.active
    url = 'http://www.lagou.com/gongsi/'
    area_name_list = get_city_name_list(url)
    for area_name in area_name_list:
        wb.create_sheet(title = area_name)
    file_name = fileName+'.xlsx'
    wb.save(file_name)
    areaId_list = get_cityId_list(url)
    for areaId in areaId_list:
        company_list = get_company_list(areaId)
        print('正在爬取----->****'+company_list[0][0]+'****公司列表')
        wb1 = load_workbook(file_name)
        ws = wb1.get_sheet_by_name(company_list[0][0])
        ws.append(['城市名称','城市得分','公司期望','公司ID','公司标签','公司Logo','发展阶段','企业名称','企业位置','企业简称','','注册时间','财务状况','行业','在招职位','其他标签','简历处理率'])
        for company in company_list:
            ws.append([company[0],str(company[1]),company[2],str(company[3]),company[4],company[5],company[6],company[7],company[8],company[9],company[10],company[11],company[12],company[13],company[14],company[15]])
        wb1.save(file_name)

file_name =  input('请输入文件名称')
print(str(datetime.datetime.now()))
write_file(file_name)
print(str(datetime.datetime.now()))
